from openpyxl import Workbook, load_workbook

wb = load_workbook('C:/Users/User/Desktop/bull/back/files/O/l_e11dep onyx.xlsx')
ws = wb.active

ws['A1'].value = "codigo"
ws['B1'].value = "descripcion"
ws['C1'].value = "tipo"
ws['D1'].value = "cant_A"
ws['E1'].value = "cant_B"

# ws.delete_rows(7084)

wb.save('C:/Users/User/Desktop/bull/back/files/O/l_e11dep onyx.xlsx')